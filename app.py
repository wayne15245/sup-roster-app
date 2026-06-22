import streamlit as st
import pandas as pd
import openpyxl
import re
from openpyxl.styles import PatternFill

# ===== UI =====
st.set_page_config(page_title="SUP Manpower Calculator", layout="wide")
st.title("📊 SUP 人手統計系統（Final Pro Max）")

uploaded_file = st.file_uploader("請上傳 Roster Excel", type=["xlsm", "xlsx"])

# ===== 設定 =====
TARGET_SHEETS = [1, 2, 3, 4, 5, 6]
SUMMARY_COLS = ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]

TIME_BLOCKS = {
    4: [5, 11],
    18: [19, 25],
    32: [33, 39],
    46: [47, 53]
}

# ===== 時間抽取 =====
TIME_PATTERN = re.compile(r'(\d{4}-\d{4})')

def extract_time(val):
    if not val:
        return None
    match = TIME_PATTERN.search(str(val))
    return match.group(1) if match else None


# ===== 防假SUP =====
def is_valid_sup(val):
    if val is None:
        return False
    val = str(val).strip()
    if val == "" or val == "0":
        return False
    if val.replace(" ", "") == "":
        return False
    return True


# ===== 分類（取代SUMMARY_MAP ✅）=====
def get_shift_group(time_str):
    try:
        start = int(time_str.split("-")[0])

        if 700 <= start <= 959:
            return "0730-1930"   # 黃
        elif 1000 <= start <= 1459:
            return "1200-0100"   # 橙
        elif 1500 <= start <= 2059:
            return "1530-0530"   # 綠
        else:
            return "2130-0930"   # 藍
    except:
        return None


# ===== Excel 顏色 =====
def get_excel_fill(group):
    if group == "0730-1930":
        return PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    elif group == "1200-0100":
        return PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    elif group == "1530-0530":
        return PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    elif group == "2130-0930":
        return PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
    elif group == "TOTAL":
        return PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    elif group == "Daily_Total":
        return PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    return None


# ===== 主流程 =====
if uploaded_file:

    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    final_result = {}

    # =========================
    # ✅ 第2–7頁
    # =========================
    for sheet_index in TARGET_SHEETS:
        sheet = wb.worksheets[sheet_index]

        for time_row, sup_rows in TIME_BLOCKS.items():
            for col_idx in range(2, 9):

                time_value = extract_time(sheet.cell(row=time_row, column=col_idx).value)
                if not time_value:
                    continue

                if time_value not in final_result:
                    final_result[time_value] = [0]*7

                count = 0
                for sup_row in sup_rows:
                    if is_valid_sup(sheet.cell(row=sup_row, column=10).value):
                        count += 1

                final_result[time_value][col_idx-2] += count

    # =========================
    # ✅ 第8頁
    # =========================
    try:
        sheet = wb.worksheets[7]
        for row in range(5, 18):
            if not is_valid_sup(sheet.cell(row=row, column=10).value):
                continue
            for col_idx in range(2, 9):
                time_value = extract_time(sheet.cell(row=row, column=col_idx).value)
                if not time_value:
                    continue
                if time_value not in final_result:
                    final_result[time_value] = [0]*7
                final_result[time_value][col_idx-2] += 1
    except:
        st.warning("⚠️ 第8頁讀取失敗")

    # =========================
    # ✅ 第9頁
    # =========================
    try:
        sheet = wb.worksheets[8]
        for row in range(4, 35):
            if not is_valid_sup(sheet.cell(row=row, column=10).value):
                continue
            for col_idx in range(2, 9):
                time_value = extract_time(sheet.cell(row=row, column=col_idx).value)
                if not time_value:
                    continue
                if time_value not in final_result:
                    final_result[time_value] = [0]*7
                final_result[time_value][col_idx-2] += 1
    except:
        st.warning("⚠️ 第9頁讀取失敗")


    # =========================
    # ✅ DEBUG
    # =========================
    if final_result:
        st.json(final_result)


    # =========================
    # ✅ DataFrame
    # =========================
    if final_result:

        df = pd.DataFrame.from_dict(
            final_result, orient="index", columns=SUMMARY_COLS
        )

        df = df.reset_index()
        df.columns = ["Time"] + SUMMARY_COLS
        df = df.sort_values(by="Time")

        df["Total"] = df[SUMMARY_COLS].sum(axis=1)


        # =========================
        # ✅ 顏色分類 SUMMARY
        # =========================
        summary_data = {
            "0730-1930": [0]*7,
            "1200-0100": [0]*7,
            "1530-0530": [0]*7,
            "2130-0930": [0]*7
        }

        for t, vals in final_result.items():
            group = get_shift_group(t)
            if not group:
                continue
            for i in range(7):
                summary_data[group][i] += vals[i]

        summary_rows = []
        for group, vals in summary_data.items():
            row = {"Time": group}
            for i, col in enumerate(SUMMARY_COLS):
                row[col] = vals[i]
            row["Total"] = sum(vals)
            summary_rows.append(row)

        df = pd.concat([df, pd.DataFrame(summary_rows)], ignore_index=True)


        # =========================
        # ✅ TOTAL（紅）
        # =========================
        total_row = {"Time": "TOTAL"}
        for col in SUMMARY_COLS:
            total_row[col] = sum(v for v in df[df["Time"] != "TOTAL"][col])
        total_row["Total"] = sum(total_row[col] for col in SUMMARY_COLS)
        df.loc[len(df)] = total_row


        # =========================
        # ✅ Daily_Total（橙）
        # =========================
        daily_total = {"Time": "Daily_Total"}
        for i, col in enumerate(SUMMARY_COLS):
            daily_total[col] = sum(v[i] for v in final_result.values())
        daily_total["Total"] = sum(daily_total[col] for col in SUMMARY_COLS)
        df.loc[len(df)] = daily_total


        # =========================
        # ✅ 最少人 shift
        # =========================
        filtered_df = df[
            ~df["Time"].isin(["TOTAL","Daily_Total","0730-1930","1200-0100","1530-0530","2130-0930"])
        ]

        min_row = filtered_df.loc[filtered_df["Total"].idxmin()]

        st.subheader("📉 最少人手時段")
        st.write(f"{min_row['Time']} ➜ {min_row['Total']} 人")


        # =========================
        # ✅ UI
        # =========================
        st.success("✅ 計算完成")
        st.dataframe(df, use_container_width=True)
        st.bar_chart(df.set_index("Time")[["Total"]])


        # =========================
        # ✅ Excel輸出 + 顏色
        # =========================
        output_file = "SUP_Result.xlsx"

        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Result")
            ws = writer.sheets["Result"]

            for row_idx in range(2, len(df)+2):
                time_val = ws.cell(row=row_idx, column=1).value
                fill = get_excel_fill(time_val)

                if fill:
                    for col in range(1, 10):
                        ws.cell(row=row_idx, column=col).fill = fill

        with open(output_file, "rb") as f:
            st.download_button("⬇️ 下載 Excel", f, file_name=output_file)

    else:
        st.error("❌ 無數據")

else:
    st.info("請上載Excel")